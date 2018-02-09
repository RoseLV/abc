import openpyxl
import sqlite3
import matplotlib.pyplot as plt


# Open the World Temperature workbook
wb = openpyxl.load_workbook("World Temperature.xlsx")
# Create another sheet called "Comparison"
anotherSheet = wb.create_sheet(title='Comparison')

print(wb.get_sheet_names())

# Calculate mean yearly temperature of Australian states (using the temperature by state table in your database)
conn = sqlite3.connect('temperature.db')
sql = conn.cursor()
sql.execute("""select sum(AverageTemperature) as sumTemp, count(AverageTemperature) as countTemp,
               state, strftime('%Y', date) as year from GlobalTemperatureByState
               where country = "Australia"
               group by state, year
               order by state, year
               """)
result = sql.fetchall()

# write to worksheet
# anotherSheet = wb.active # this will write the data to the current active sheet
for i in range(len(result)):
    anotherSheet.cell(row=i + 2, column=1).value = result[i][0] / result[i][1]
    anotherSheet.cell(row=i + 2, column=2).value = result[i][3]
    anotherSheet.cell(row=i + 2, column=3).value = result[i][2]

# column title
anotherSheet.cell(row=1, column=1).value = "Avgtemp"
anotherSheet.cell(row=1, column=2).value = "Year"
anotherSheet.cell(row=1, column=3).value = "Aus-State"


stateMap = dict()
for row in result:
    sumTemp, months, state, year = row
    avgTemp = sumTemp/months

    if state not in stateMap:
        stateMap[state] = ([], []) #tuple(list(), list())

    avgTemp, forYear = stateMap[state]
    avgTemp.append(sumTemp / months)
    forYear.append(year)



# Calculate mean yearly temperature of Australia using table 'GlobalTemperatureByCountry'
sql.execute("""select sum(AverageTemperature) as sumTemp, count(AverageTemperature) as countTemp,
               strftime('%Y', date) as year from GlobalTemperatureByCountry
               where country = "Australia"
               group by year
               """)
result = sql.fetchall()

for row in result:
    sumTemp, months, year = row
    avgTemp = sumTemp / months
    if "Australia" not in stateMap:
        stateMap["Australia"] = ([], [])  # tuple(list(), list())

    avgTemp, forYear = stateMap["Australia"]
    avgTemp.append(sumTemp / months)
    forYear.append(year)
print(stateMap)

# write country data to worksheet
for i in range(len(result)):
#for state, (avgTemp, year) in stateMap.items():
    anotherSheet.cell(row=i + 1347, column=1).value = result[i][0] / result[i][1]
    anotherSheet.cell(row=i + 1347, column=2).value = result[i][2]
    anotherSheet.cell(row=i + 1347, column=3).value = 'Australia'
wb.save('World Temperature.xlsx')

# Plot differences between each state & national data for each year.
plt.title('Yearly Average Temperature Difference In Australia')
plt.xlabel('Year')
plt.ylabel('Temperature')

for state, (avgTemp, year) in stateMap.items():
    plt.plot(year, avgTemp, label=str(state))

# create the legend in a box to tell the lines
keys = list(stateMap.keys())
plt.legend(keys, loc=0)
plt.show()


conn.commit()
conn.close()
