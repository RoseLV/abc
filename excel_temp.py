import openpyxl
import sqlite3
import os
import matplotlib.pyplot as plt

# Delete excel if created
excel_filename = 'World Temperature.xlsx'
if os.path.exists(excel_filename):
    os.remove(excel_filename)

# Create the 'World Temperature.xlsx' workbook, and 'Temperature by city' sheet
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Temperature by city'
wb.save('World Temperature.xlsx')

# Connect database
conn = sqlite3.connect('temperature.db')
sql = conn.cursor()

# Create small table : ChineseCityYearlyAvgTemperature
sql.execute('drop table if exists ChineseCityYearlyAvgTemperature')
sql.execute('create table ChineseCityYearlyAvgTemperature(date Date, temp REAL, city text)')

# Import data to small table from big table : GlobalTemperatureByMajorCity
sql.execute("""select AverageTemperature, city, date from GlobalTemperatureByMajorCity
                where country = 'China' order by date
            """)

result = sql.fetchall()
# Handle mission data
for row in result:
    if row[0] != 'None':
        temp, city, date = row
        # insert some data from big table to small table
    sql.execute('insert into ChineseCityYearlyAvgTemperature values("{date}", "{temp}", "{city}");'.format(
        date=date, temp=temp, city=city))

sql.execute("""select sum(temp) as totalAvgTemp, count(temp) as tempDenom, 
                city, strftime('%Y', date) as year 
                from ChineseCityYearlyAvgTemperature
                group by city, year
                order by city, year
            """)

result = sql.fetchall()
cityMap = dict()

for row in result:
    sumTemp, months, city, year = row

    if city not in cityMap:
        cityMap[city] = ([], [])

    avgTemp, forYear = cityMap[city]
    avgTemp.append(sumTemp / months)
    forYear.append(year)

    #cityMap[city][0].append(sumTemp/countMonth)
    #cityMap[city][1].append(year)


plt.title('Yearly Average Temperature Change In Chinese Major Cities')
plt.xlabel('Year')
plt.ylabel('Temperature')

#for city in cityMap:
    #l1 = plt.plot(cityMap[city][1], cityMap[city][0])
for city, (avgTemp, year) in cityMap.items():
    plt.plot(year, avgTemp, label=str(city))

# create the legend in a box to tell the lines
keys = list(cityMap.keys())
plt.legend(keys, loc=0)
plt.show()

# Write into excel
wb = openpyxl.load_workbook('World Temperature.xlsx')
sheet = wb.active

"""for i in range(len(result)):
    AvgTemp = result[i][0]/result[i][1]
    city = result[i][2]
    year = result[i][3]
    print('{:.2f}'.format(AvgTemp),city,year)"""

for i in range(len(result)):
    sheet.cell(row=i+2, column=1).value = result[i][0]/result[i][1]
    sheet.cell(row=i+2, column=2).value = result[i][2]
    sheet.cell(row=i+2, column=3).value = result[i][3]
# column title
sheet.cell(row=1, column=1).value = "Avgtemp"
sheet.cell(row=1, column=2).value = "cityName"
sheet.cell(row=1, column=3).value = "Year"

wb.save('World Temperature.xlsx')
conn.commit()
conn.close()