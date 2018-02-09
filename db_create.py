import openpyxl
import sqlite3
import os

# drop DB if exists
db_filename = 'temperature.db'
if os.path.exists(db_filename):
    os.remove(db_filename)

# connect db
conn = sqlite3.connect('temperature.db')
sql = conn.cursor()

# create table GlobalTemperatureByCountry and drop if table already exists
sql.execute("drop table if exists GlobalTemperatureByCountry")
sql.execute('''create table GlobalTemperatureByCountry
(date text, AverageTemperature REAL, AverageTemperatureUncertainty REAL, country text)''')

# Open the excel file 'GlobalLandTemperaturesByCountry.xlsx'
wb = openpyxl.load_workbook('GlobalLandTemperaturesByCountry.xlsx')
sheet = wb.active
maxRow = sheet.max_row

# Import the data from the excel files to the corresponding tables in the database
for i in range(2, maxRow+1):
    date = str(sheet.cell(row=i, column=1).value)
    avgTemp = sheet.cell(row=i, column=2).value
    avgTempUncertainty = sheet.cell(row=i, column=3).value
    country = str(sheet.cell(row=i, column=4).value)
    #print(date, avgTemp, avgTempUncertainty, country)
    sql.execute('insert into GlobalTemperatureByCountry values("{date}", "{temp}", "{tempUncertainty}", "{country}");'.format(
        date=date, temp=avgTemp, tempUncertainty=avgTempUncertainty, country=country
    ))
    #sql.execute(""{date}" , "{temp}" , "{tempUncertainty}" , "{country}");")
conn.commit()


# create table GlobalTemperatureByState and drop table if exists
sql.execute("""drop table if exists GlobalTemperatureByState""")
sql.execute("""create table GlobalTemperatureByState
(date text, AverageTemperature REAL, AverageTemperatureUncertainty REAL, state text, country text)""")

# load data from workbook GlobalLandTemperaturesByState.xlsx
wb = openpyxl.load_workbook('GlobalLandTemperaturesByState.xlsx')
sheet = wb.active
maxRow = sheet.max_row
for i in range(2, maxRow+1):
    date = str(sheet.cell(row=i, column=1).value)
    avgTemp = sheet.cell(row=i, column=2).value
    avgTempUncertainty = sheet.cell(row=i, column=3).value
    state = str(sheet.cell(row=i, column=4).value)
    country = str(sheet.cell(row=i, column=5).value)
    sql.execute('insert into GlobalTemperatureByState values("{date}", "{temp}", "{tempUncertainty}","{state}", "{country}");'.format(
        date=date, temp=avgTemp, tempUncertainty=avgTempUncertainty, state=state, country=country
    ))
conn.commit()


# create table GlobalTemperatureByMajorCity and drop table if exists
sql.execute("""drop table if exists GlobalTemperatureByMajorCity;""")
sql.execute('''create table GlobalTemperatureByMajorCity
(date Date, AverageTemperature REAL, AverageTemperatureUncertainty REAL, 
city text, country text, latitude text, longitude text)''')

# load data from workbook GlobalLandTemperaturesByMajorCity.xlsx
wb = openpyxl.load_workbook('GlobalLandTemperaturesByMajorCity.xlsx')
sheet = wb.active
maxRow = sheet.max_row

for i in range(2, maxRow+1):
    #for j in range(1,8):
    date = str(sheet.cell(row=i, column=1).value)
    avgTemp = sheet.cell(row=i, column=2).value
    avgTempUncertainty = sheet.cell(row=i, column=3).value
    city = str(sheet.cell(row=i, column=4).value)
    country = str(sheet.cell(row=i, column=5).value)
    latitude = str(sheet.cell(row=i, column=6).value)
    longitude = str(sheet.cell(row=i, column=7).value)
    #print(date, avgTemp, avgTempUncertainty, city, country, latitude, longitude)
    sql.execute('insert into GlobalTemperatureByMajorCity values("{date}", "{temp}", "{tempUncertainty}", "{city}", "{country}","{latitude}","{longitude}");'.format(
        date=date, temp=avgTemp, tempUncertainty=avgTempUncertainty, city=city, country=country, latitude=latitude, longitude=longitude
    ))
conn.commit()
conn.close()
